"""
src/reports/build_peer_comparison_xlsx.py - Module 4, Day 20:
peer_comparison.xlsx - one sheet per peer group, percentile colour-coded,
benchmark row highlighted gold, median summary row.
"""
import sqlite3, sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font

DB_PATH = Path(__file__).resolve().parent.parent.parent / "data" / "nifty100.db"
OUT_PATH = Path(__file__).resolve().parent.parent.parent / "output" / "peer_comparison.xlsx"

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GOLD = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
HEADER_FONT = Font(bold=True)
BOLD = Font(bold=True)


def build():
    conn = sqlite3.connect(DB_PATH)
    pp = pd.read_sql("SELECT * FROM peer_percentiles", conn)
    companies = pd.read_sql("SELECT id, company_name FROM companies", conn)
    benchmarks = pd.read_sql("SELECT peer_group_name, company_id FROM peer_groups WHERE is_benchmark = 1", conn)
    benchmark_map = dict(zip(benchmarks["peer_group_name"], benchmarks["company_id"]))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for group_name in sorted(pp["peer_group_name"].unique()):
        group_data = pp[pp["peer_group_name"] == group_name]
        metrics = sorted(group_data["metric"].unique())
        members = sorted(group_data["company_id"].unique())

        ws = wb.create_sheet(group_name[:31])
        header = ["company_id", "company_name"] + metrics
        ws.append(header)
        for cell in ws[1]:
            cell.font = HEADER_FONT

        for cid in members:
            name = companies.loc[companies["id"] == cid, "company_name"].values
            row_vals = [cid, name[0] if len(name) else ""]
            pct_by_metric = {}
            for m in metrics:
                rec = group_data[(group_data["company_id"] == cid) & (group_data["metric"] == m)]
                val = rec["value"].values[0] if len(rec) else None
                pct = rec["percentile_rank"].values[0] if len(rec) else None
                row_vals.append(val)
                pct_by_metric[m] = pct
            ws.append(row_vals)
            r = ws.max_row
            is_benchmark = benchmark_map.get(group_name) == cid
            for j, m in enumerate(metrics, start=3):
                pct = pct_by_metric[m]
                if pct is None:
                    continue
                cell = ws.cell(row=r, column=j)
                if pct >= 0.75:
                    cell.fill = GREEN
                elif pct <= 0.25:
                    cell.fill = RED
                else:
                    cell.fill = YELLOW
            if is_benchmark:
                for j in range(1, len(header) + 1):
                    ws.cell(row=r, column=j).fill = GOLD

        median_row = ["MEDIAN", ""]
        for m in metrics:
            vals = group_data[group_data["metric"] == m]["value"].dropna()
            median_row.append(vals.median() if len(vals) else None)
        ws.append(median_row)
        for cell in ws[ws.max_row]:
            cell.font = BOLD

        for col_cells in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 25)

    wb.save(OUT_PATH)
    conn.close()
    return wb.sheetnames


if __name__ == "__main__":
    sheets = build()
    print(f"Saved. Sheets: {len(sheets)}")
    for s in sheets:
        print(f"  {s}")