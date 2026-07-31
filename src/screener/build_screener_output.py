import sqlite3, sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from screener.engine import build_universe, load_config, run_preset
from analytics.composite_score import compute_composite_scores
import openpyxl
from openpyxl.styles import PatternFill, Font

DB_PATH = Path(__file__).resolve().parent.parent.parent / "data" / "nifty100.db"

conn = sqlite3.connect(DB_PATH)
config = load_config()
universe = build_universe(conn)
universe = compute_composite_scores(universe, conn)

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True)

wb = openpyxl.Workbook()
wb.remove(wb.active)

display_cols = ["company_id", "company_name", "return_on_equity_pct", "debt_to_equity",
                 "free_cash_flow_cr", "revenue_cagr_5yr", "pat_cagr_5yr",
                 "pe_ratio", "pb_ratio", "dividend_yield_pct",
                 "composite_quality_score_sector"]

for preset_name, preset_cfg in config["presets"].items():
    result = run_preset(preset_name, universe, config, conn)
    ws = wb.create_sheet(preset_cfg["label"][:31])
    ws.append([c.replace("_", " ").title() for c in display_cols])
    for cell in ws[1]:
        cell.font = HEADER_FONT

    filter_cols = {config["metrics"][k]["column"] for k in preset_cfg.get("filters", {}) if k in config["metrics"]}
    for _, row in result.iterrows():
        ws.append([row.get(c) for c in display_cols])
        for j, c in enumerate(display_cols, start=1):
            if c in filter_cols:
                ws.cell(row=ws.max_row, column=j).fill = GREEN

    for col_cells in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 30)

wb.save(str(Path(__file__).resolve().parent.parent.parent / "output" / "screener_output.xlsx"))
print("Saved. Sheets:", wb.sheetnames)
for name in wb.sheetnames:
    print(f"  {name}: {wb[name].max_row - 1} companies")
conn.close()